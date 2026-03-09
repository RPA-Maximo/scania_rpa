"""
仓库和仓位信息同步服务

功能：
- 从 Maximo 同步仓库（storeroom）信息到 warehouse 表
- 从 Maximo 库存数据提取仓位（bin）并写入 warehouse_bin 表
- 导出仓库/仓位/关联关系 Excel
- 导入 Excel 维护仓库-仓位关联关系
"""
import io
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.fetcher.warehouse_fetcher import fetch_warehouses, fetch_bins_from_inventory
from src.utils.db import get_connection, generate_id


# ── 表结构初始化 ───────────────────────────────────────────────────────────────


def _add_col(cursor, table: str, col: str, definition: str):
    cursor.execute(f"SHOW COLUMNS FROM `{table}` LIKE %s", (col,))
    if cursor.fetchone() is None:
        cursor.execute(f"ALTER TABLE `{table}` ADD COLUMN `{col}` {definition}")


def ensure_warehouse_tables(conn):
    """创建 warehouse 和 warehouse_bin 表（已存在则跳过）"""
    cursor = conn.cursor()
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `warehouse` (
                `id`            BIGINT        NOT NULL PRIMARY KEY COMMENT '主键',
                `code`          VARCHAR(50)   NOT NULL COMMENT '仓库编号(Maximo location)',
                `name`          VARCHAR(200)  NULL     COMMENT '仓库名称(description)',
                `site`          VARCHAR(50)   NULL     COMMENT '地点(siteid)',
                `org`           VARCHAR(50)   NULL     COMMENT '组织(orgid)',
                `location_type` VARCHAR(30)   NULL     COMMENT '位置类型',
                `status`        VARCHAR(20)   NULL     COMMENT '状态',
                `sync_time`     DATETIME      NULL     COMMENT '最近同步时间',
                `create_time`   DATETIME      NOT NULL DEFAULT CURRENT_TIMESTAMP,
                `update_time`   DATETIME      NULL     ON UPDATE CURRENT_TIMESTAMP,
                `del_flag`      TINYINT       NOT NULL DEFAULT 0,
                UNIQUE KEY `uq_warehouse_code` (`code`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='仓库主数据表'
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `warehouse_bin` (
                `id`           BIGINT        NOT NULL PRIMARY KEY COMMENT '主键',
                `warehouse`    VARCHAR(50)   NOT NULL COMMENT '仓库编号',
                `bin_code`     VARCHAR(100)  NOT NULL COMMENT '仓位编号',
                `bin_name`     VARCHAR(200)  NULL     COMMENT '仓位名称',
                `site`         VARCHAR(50)   NULL     COMMENT '地点(siteid)',
                `remark`       VARCHAR(500)  NULL     COMMENT '备注',
                `sync_source`  VARCHAR(20)   NOT NULL DEFAULT 'maximo' COMMENT '数据来源(maximo/import)',
                `create_time`  DATETIME      NOT NULL DEFAULT CURRENT_TIMESTAMP,
                `update_time`  DATETIME      NULL     ON UPDATE CURRENT_TIMESTAMP,
                `del_flag`     TINYINT       NOT NULL DEFAULT 0,
                UNIQUE KEY `uq_wh_bin` (`warehouse`, `bin_code`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='仓库-仓位关联表'
        """)
        conn.commit()
    except Exception as e:
        print(f"[WARN] ensure_warehouse_tables: {e}")
    finally:
        cursor.close()


# ── 仓库同步 ──────────────────────────────────────────────────────────────────


def sync_warehouses(
    site_id: Optional[str] = None,
    max_pages: int = 20,
    page_size: int = 100,
) -> Dict[str, int]:
    """
    从 Maximo MXAPILOCATION 同步仓库主数据到 warehouse 表

    Returns:
        {'inserted': N, 'updated': N, 'skipped': N}
    """
    raw_items = fetch_warehouses(
        site_id=site_id,
        max_pages=max_pages,
        page_size=page_size,
    )

    if not raw_items:
        print("[WARN] 未获取到任何仓库数据")
        return {"inserted": 0, "updated": 0, "skipped": 0}

    conn = get_connection()
    try:
        ensure_warehouse_tables(conn)
        cursor = conn.cursor(dictionary=True)
        stats = {"inserted": 0, "updated": 0, "skipped": 0}
        now = datetime.now()

        for item in raw_items:
            code = str(item.get("location") or "").strip()
            if not code:
                stats["skipped"] += 1
                continue

            name   = str(item.get("description") or "")[:200]
            site   = str(item.get("siteid") or "")[:50]
            org    = str(item.get("orgid") or "")[:50]
            ltype  = str(item.get("type") or "")[:30]
            status = str(item.get("status") or "")[:20]

            cursor.execute("SELECT id FROM warehouse WHERE code=%s", (code,))
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    """UPDATE warehouse SET
                        name=%s, site=%s, org=%s, location_type=%s,
                        status=%s, sync_time=%s, del_flag=0, update_time=%s
                       WHERE id=%s""",
                    (name, site, org, ltype, status, now, now, existing["id"]),
                )
                stats["updated"] += 1
            else:
                cursor.execute(
                    """INSERT INTO warehouse
                        (id, code, name, site, org, location_type,
                         status, sync_time, create_time, del_flag)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,0)""",
                    (generate_id(), code, name, site, org, ltype, status, now, now),
                )
                stats["inserted"] += 1

        conn.commit()
        print(f"[OK] 仓库同步完成: {stats}")
        return stats

    except Exception as e:
        conn.rollback()
        print(f"[ERROR] 仓库同步失败: {e}")
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        conn.close()


# ── 仓位同步 ──────────────────────────────────────────────────────────────────


def sync_warehouse_bins(
    warehouse: Optional[str] = None,
    site_id: Optional[str] = None,
    max_pages: int = 50,
    page_size: int = 100,
) -> Dict[str, int]:
    """
    从 Maximo 库存数据提取仓位信息，写入 warehouse_bin 表

    Returns:
        {'inserted': N, 'updated': N, 'skipped': N}
    """
    bin_rows = fetch_bins_from_inventory(
        warehouse=warehouse,
        site_id=site_id,
        max_pages=max_pages,
        page_size=page_size,
    )

    if not bin_rows:
        print("[WARN] 未获取到任何仓位数据")
        return {"inserted": 0, "updated": 0, "skipped": 0}

    conn = get_connection()
    try:
        ensure_warehouse_tables(conn)
        cursor = conn.cursor(dictionary=True)
        stats = {"inserted": 0, "updated": 0, "skipped": 0}
        now = datetime.now()

        for row in bin_rows:
            wh  = row["warehouse"]
            bin = row["bin_code"]
            site = row.get("site", "")

            if not wh or not bin:
                stats["skipped"] += 1
                continue

            cursor.execute(
                "SELECT id FROM warehouse_bin WHERE warehouse=%s AND bin_code=%s",
                (wh, bin),
            )
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    """UPDATE warehouse_bin SET
                        site=%s, del_flag=0, update_time=%s
                       WHERE id=%s""",
                    (site, now, existing["id"]),
                )
                stats["updated"] += 1
            else:
                cursor.execute(
                    """INSERT INTO warehouse_bin
                        (id, warehouse, bin_code, site, sync_source, create_time, del_flag)
                       VALUES (%s,%s,%s,%s,'maximo',%s,0)""",
                    (generate_id(), wh, bin, site, now),
                )
                stats["inserted"] += 1

        conn.commit()
        print(f"[OK] 仓位同步完成: {stats}")
        return stats

    except Exception as e:
        conn.rollback()
        print(f"[ERROR] 仓位同步失败: {e}")
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        conn.close()


# ── Excel 导出 ────────────────────────────────────────────────────────────────


def export_warehouse_excel(
    include_bins: bool = True,
    warehouse: Optional[str] = None,
) -> bytes:
    """
    导出仓库+仓位信息及关联关系为 Excel（多个工作表）

    Args:
        include_bins: 是否包含仓位关联工作表
        warehouse:    仓库过滤；None=全部

    Returns:
        Excel 文件字节内容
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # 仓库数据
        wh_where = ["del_flag=0"]
        wh_params = []
        if warehouse:
            wh_where.append("code=%s")
            wh_params.append(warehouse)
        cursor.execute(
            f"""SELECT code, name, site, org, location_type, status, sync_time
                FROM warehouse WHERE {' AND '.join(wh_where)}
                ORDER BY code""",
            wh_params,
        )
        wh_rows = cursor.fetchall()

        # 仓位数据
        bin_rows = []
        if include_bins:
            bin_where = ["del_flag=0"]
            bin_params = []
            if warehouse:
                bin_where.append("warehouse=%s")
                bin_params.append(warehouse)
            cursor.execute(
                f"""SELECT warehouse, bin_code, bin_name, site, remark, sync_source
                    FROM warehouse_bin WHERE {' AND '.join(bin_where)}
                    ORDER BY warehouse, bin_code""",
                bin_params,
            )
            bin_rows = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    def _write_sheet(ws, cols, rows, col_widths):
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        ws.row_dimensions[1].height = 22
        for r in rows:
            ws.append([r.get(c) if not isinstance(r.get(c), datetime) else str(r.get(c))[:19] for c in cols])
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w

    # Sheet1: 仓库
    ws1 = wb.active
    ws1.title = "仓库信息"
    _write_sheet(
        ws1,
        ["code", "name", "site", "org", "location_type", "status", "sync_time"],
        wh_rows,
        {"A": 15, "B": 35, "C": 12, "D": 12, "E": 15, "F": 12, "G": 20},
    )
    # 改用中文列头
    ws1["A1"] = "仓库编号"
    ws1["B1"] = "仓库名称"
    ws1["C1"] = "地点(Site)"
    ws1["D1"] = "组织(Org)"
    ws1["E1"] = "类型"
    ws1["F1"] = "状态"
    ws1["G1"] = "同步时间"

    # Sheet2: 仓库-仓位关联
    if include_bins:
        ws2 = wb.create_sheet("仓库-仓位关联")
        _write_sheet(
            ws2,
            ["warehouse", "bin_code", "bin_name", "site", "remark", "sync_source"],
            bin_rows,
            {"A": 15, "B": 20, "C": 25, "D": 12, "E": 30, "F": 12},
        )
        ws2["A1"] = "仓库编号"
        ws2["B1"] = "仓位编号"
        ws2["C1"] = "仓位名称"
        ws2["D1"] = "地点(Site)"
        ws2["E1"] = "备注"
        ws2["F1"] = "数据来源"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Excel 导入（仓库-仓位关联） ───────────────────────────────────────────────


def import_warehouse_bins_excel(file_bytes: bytes) -> Dict:
    """
    从 Excel 导入仓库-仓位关联关系到 warehouse_bin 表

    Excel 格式（至少含以下列，顺序不限）：
      - 仓库编号 / warehouse
      - 仓位编号 / bin_code
      - 仓位名称 / bin_name  （可选）
      - 备注     / remark    （可选）

    Returns:
        {'inserted': N, 'updated': N, 'skipped': N, 'warnings': [...]}
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return {"inserted": 0, "updated": 0, "skipped": 0,
                "warnings": ["Excel 无数据行"]}

    # 解析列索引（兼容中英文列头）
    WH_ALIASES   = {"仓库编号", "warehouse", "storeloc", "仓库"}
    BIN_ALIASES  = {"仓位编号", "bin_code", "binnum", "仓位"}
    NAME_ALIASES = {"仓位名称", "bin_name", "仓位名"}
    REM_ALIASES  = {"备注", "remark"}

    header_row = [str(c or "").strip().lower() for c in all_rows[0]]
    wh_idx = bin_idx = name_idx = rem_idx = None
    for i, h in enumerate(header_row):
        if h in {a.lower() for a in WH_ALIASES}:
            wh_idx = i
        elif h in {a.lower() for a in BIN_ALIASES}:
            bin_idx = i
        elif h in {a.lower() for a in NAME_ALIASES}:
            name_idx = i
        elif h in {a.lower() for a in REM_ALIASES}:
            rem_idx = i

    if wh_idx is None or bin_idx is None:
        return {"inserted": 0, "updated": 0, "skipped": 0,
                "warnings": ["未找到"仓库编号"或"仓位编号"列"]}

    conn = get_connection()
    try:
        ensure_warehouse_tables(conn)
        cursor = conn.cursor(dictionary=True)
        stats = {"inserted": 0, "updated": 0, "skipped": 0, "warnings": []}
        now = datetime.now()

        for row_i, row in enumerate(all_rows[1:], start=2):
            wh  = str(row[wh_idx]  or "").strip()
            bin = str(row[bin_idx] or "").strip()
            if not wh or not bin:
                stats["skipped"] += 1
                continue

            name   = str(row[name_idx] or "").strip() if name_idx is not None else ""
            remark = str(row[rem_idx]  or "").strip() if rem_idx  is not None else ""

            cursor.execute(
                "SELECT id FROM warehouse_bin WHERE warehouse=%s AND bin_code=%s",
                (wh, bin),
            )
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    """UPDATE warehouse_bin SET
                        bin_name=%s, remark=%s, sync_source='import',
                        del_flag=0, update_time=%s
                       WHERE id=%s""",
                    (name or None, remark or None, now, existing["id"]),
                )
                stats["updated"] += 1
            else:
                cursor.execute(
                    """INSERT INTO warehouse_bin
                        (id, warehouse, bin_code, bin_name, remark,
                         sync_source, create_time, del_flag)
                       VALUES (%s,%s,%s,%s,%s,'import',%s,0)""",
                    (generate_id(), wh, bin, name or None, remark or None, now),
                )
                stats["inserted"] += 1

        conn.commit()
        return stats

    except Exception as e:
        conn.rollback()
        raise RuntimeError(f"导入失败: {e}") from e
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        conn.close()
