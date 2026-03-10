"""
物料单价同步服务

功能：
- 从 Maximo MXAPIINVENTORY 同步物料单价（LIFO/FIFO unitcost）到 material 表
- 导出库存报表（物料+库存数量+单价+货值）为 Excel
"""
import io
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.fetcher.invcost_fetcher import fetch_invcost
from src.utils.db import get_connection


# ── 表字段初始化 ──────────────────────────────────────────────────────────────


def _add_col(cursor, table: str, col: str, definition: str):
    cursor.execute(f"SHOW COLUMNS FROM `{table}` LIKE %s", (col,))
    if cursor.fetchone() is None:
        cursor.execute(f"ALTER TABLE `{table}` ADD COLUMN `{col}` {definition}")


def ensure_material_cost_columns(conn):
    """在 material 表中添加单价相关字段（不存在则添加）"""
    cursor = conn.cursor()
    try:
        _add_col(cursor, "material", "unit_cost",
                 "DECIMAL(18,4) NULL COMMENT '物料单价(来自Maximo invcost)'")
        _add_col(cursor, "material", "avg_cost",
                 "DECIMAL(18,4) NULL COMMENT '物料平均成本(avgcost)'")
        _add_col(cursor, "material", "last_cost",
                 "DECIMAL(18,4) NULL COMMENT '物料最近成本(lastcost)'")
        _add_col(cursor, "material", "cost_date",
                 "DATETIME NULL COMMENT '成本日期(invcost costdate)'")
        _add_col(cursor, "material", "cost_sync_time",
                 "DATETIME NULL COMMENT '单价最近同步时间'")
        conn.commit()
    except Exception as e:
        print(f"[WARN] ensure_material_cost_columns: {e}")
    finally:
        cursor.close()


# ── 核心同步 ──────────────────────────────────────────────────────────────────


def sync_invcost(
    item_numbers: Optional[List[str]] = None,
    warehouse: Optional[str] = None,
    max_pages: int = 100,
    page_size: int = 50,
) -> Dict[str, int]:
    """
    从 Maximo 同步物料单价到 material 表的 unit_cost 字段

    以 item_number(code) 为匹配键，更新已有物料的单价信息。
    若同一物料在多个仓库有不同单价，取最近 cost_date 的那条。

    Returns:
        {'updated': N, 'skipped': N, 'not_found': N}
    """
    raw = fetch_invcost(
        item_numbers=item_numbers,
        warehouse=warehouse,
        max_pages=max_pages,
        page_size=page_size,
    )

    if not raw:
        print("[WARN] 未获取到任何库存成本数据")
        return {"updated": 0, "skipped": 0, "not_found": 0}

    # 若同一物料有多条（不同仓库），取 unit_cost 最大值（或按需改为最新 cost_date）
    best: Dict[str, dict] = {}
    for row in raw:
        code = row["item_number"]
        if not code:
            continue
        existing = best.get(code)
        if existing is None:
            best[code] = row
        else:
            # 优先保留 cost_date 最新的
            if (row.get("cost_date") or "") > (existing.get("cost_date") or ""):
                best[code] = row

    conn = get_connection()
    try:
        ensure_material_cost_columns(conn)
        cursor = conn.cursor(dictionary=True)
        stats = {"updated": 0, "skipped": 0, "not_found": 0}
        now = datetime.now()

        for code, row in best.items():
            unit_cost  = row.get("unit_cost")
            avg_cost   = row.get("avg_cost")
            last_cost  = row.get("last_cost")
            cost_date_s = row.get("cost_date")
            cost_date  = None
            if cost_date_s:
                try:
                    cost_date = datetime.strptime(cost_date_s[:19], "%Y-%m-%dT%H:%M:%S")
                except Exception:
                    try:
                        cost_date = datetime.strptime(cost_date_s[:10], "%Y-%m-%d")
                    except Exception:
                        pass

            if unit_cost is None and avg_cost is None:
                stats["skipped"] += 1
                continue

            cursor.execute(
                "SELECT id FROM material WHERE code=%s AND del_flag=0",
                (code,),
            )
            mat = cursor.fetchone()
            if not mat:
                stats["not_found"] += 1
                continue

            cursor.execute(
                """UPDATE material SET
                    unit_cost=%s, avg_cost=%s, last_cost=%s,
                    cost_date=%s, cost_sync_time=%s
                   WHERE id=%s""",
                (unit_cost, avg_cost, last_cost, cost_date, now, mat["id"]),
            )
            stats["updated"] += 1

        conn.commit()
        print(f"[OK] 物料单价同步完成: {stats}")
        return stats

    except Exception as e:
        conn.rollback()
        print(f"[ERROR] 物料单价同步失败: {e}")
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        conn.close()


# ── 库存报表导出 ──────────────────────────────────────────────────────────────


def export_inventory_report_excel(
    warehouse: Optional[str] = None,
    item_number: Optional[str] = None,
) -> bytes:
    """
    导出库存报表 Excel

    报表内容：物料编号、物料名称、仓库、货柜、批次、库存数量、单价、货值（数量×单价）

    数据来源：bin_inventory（库存）+ material（物料主数据，含单价）
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        where  = ["b.del_flag=0"]
        params = []
        if warehouse:
            where.append("b.warehouse=%s")
            params.append(warehouse)
        if item_number:
            where.append("b.item_number LIKE %s")
            params.append(f"%{item_number}%")

        cursor.execute(
            f"""SELECT
                    b.item_number                             AS 物料编号,
                    COALESCE(m.name, '')                      AS 物料名称,
                    b.warehouse                               AS 仓库,
                    b.bin_code                                AS 货柜编号,
                    COALESCE(b.lot_number, '')                AS 批次号,
                    b.quantity                                AS 库存数量,
                    COALESCE(m.unit_cost, m.avg_cost)         AS 单价,
                    CASE
                        WHEN COALESCE(m.unit_cost, m.avg_cost) IS NOT NULL
                        THEN b.quantity * COALESCE(m.unit_cost, m.avg_cost)
                        ELSE NULL
                    END                                       AS 货值,
                    COALESCE(m.cost_date, '')                 AS 成本日期,
                    b.receipt_date                            AS 入库日期,
                    b.update_time                             AS 更新时间
                FROM bin_inventory b
                LEFT JOIN material m ON m.code = b.item_number AND m.del_flag=0
                WHERE {' AND '.join(where)}
                ORDER BY b.item_number, b.warehouse, b.bin_code""",
            params,
        )
        rows = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存报表"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    col_names = ["物料编号", "物料名称", "仓库", "货柜编号", "批次号",
                 "库存数量", "单价", "货值", "成本日期", "入库日期", "更新时间"]
    ws.append(col_names)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 22

    total_value = 0.0
    for r in rows:
        # 格式化日期
        receipt = str(r["入库日期"])[:10] if r["入库日期"] else ""
        update  = str(r["更新时间"])[:19] if r["更新时间"] else ""
        cost_dt = str(r["成本日期"])[:19] if r["成本日期"] else ""
        货值 = float(r["货值"]) if r["货值"] is not None else None
        if 货值:
            total_value += 货值

        ws.append([
            r["物料编号"],
            r["物料名称"],
            r["仓库"],
            r["货柜编号"],
            r["批次号"],
            float(r["库存数量"]) if r["库存数量"] else 0,
            float(r["单价"]) if r["单价"] is not None else None,
            round(货值, 4) if 货值 is not None else None,
            cost_dt,
            receipt,
            update,
        ])

    # 汇总行
    ws.append([])
    summary_row = ws.max_row + 1
    ws.append(["合计", "", "", "", "", "", "", round(total_value, 4), "", "", ""])
    bold_font = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.font = bold_font

    # 列宽
    for col, w in zip("ABCDEFGHIJK", [15, 30, 10, 15, 15, 12, 12, 14, 20, 12, 20]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
