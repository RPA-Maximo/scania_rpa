"""
供应商账户同步服务

功能：
- 从 Maximo MXAPICOMPANY 同步供应商编号和供应商名称到 vendor 表
- 支持全量/增量同步
- 支持导出 Excel 文件（供客户导入用）
"""
import io
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.fetcher.vendor_fetcher import fetch_vendors
from src.utils.db import get_connection, generate_id


# ── 表结构初始化 ───────────────────────────────────────────────────────────────


def _add_col(cursor, table: str, col: str, definition: str):
    cursor.execute(f"SHOW COLUMNS FROM `{table}` LIKE %s", (col,))
    if cursor.fetchone() is None:
        cursor.execute(f"ALTER TABLE `{table}` ADD COLUMN `{col}` {definition}")


def ensure_vendor_table(conn):
    """创建或更新 vendor 表"""
    cursor = conn.cursor()
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS `vendor` (
                `id`          BIGINT        NOT NULL PRIMARY KEY COMMENT '主键',
                `vendor_code` VARCHAR(50)   NOT NULL COMMENT '供应商编号(Maximo company)',
                `vendor_name` VARCHAR(200)  NULL     COMMENT '供应商名称',
                `vendor_type` VARCHAR(20)   NULL     COMMENT '供应商类型',
                `status`      VARCHAR(20)   NULL     COMMENT '状态',
                `currency`    VARCHAR(10)   NULL     COMMENT '币种',
                `sync_time`   DATETIME      NULL     COMMENT '最近同步时间',
                `create_time` DATETIME      NOT NULL DEFAULT CURRENT_TIMESTAMP,
                `update_time` DATETIME      NULL     ON UPDATE CURRENT_TIMESTAMP,
                `del_flag`    TINYINT       NOT NULL DEFAULT 0,
                UNIQUE KEY `uq_vendor_code` (`vendor_code`)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='供应商账户表'
        """)
        conn.commit()
    except Exception as e:
        print(f"[WARN] ensure_vendor_table: {e}")
    finally:
        cursor.close()


# ── 核心同步 ──────────────────────────────────────────────────────────────────


def sync_vendors(
    vendor_numbers: Optional[List[str]] = None,
    vendor_type: Optional[str] = None,
    max_pages: int = 50,
    page_size: int = 100,
) -> Dict[str, int]:
    """
    从 Maximo 同步供应商数据到 vendor 表

    Returns:
        {'inserted': N, 'updated': N, 'skipped': N}
    """
    raw_items = fetch_vendors(
        vendor_numbers=vendor_numbers,
        vendor_type=vendor_type,
        max_pages=max_pages,
        page_size=page_size,
    )

    if not raw_items:
        print("[WARN] 未获取到任何供应商数据")
        return {"inserted": 0, "updated": 0, "skipped": 0}

    conn = get_connection()
    try:
        ensure_vendor_table(conn)
        cursor = conn.cursor(dictionary=True)
        stats = {"inserted": 0, "updated": 0, "skipped": 0}
        now = datetime.now()

        for item in raw_items:
            code = str(item.get("company") or "").strip()
            if not code:
                stats["skipped"] += 1
                continue

            name  = str(item.get("name") or "")[:200]
            vtype = str(item.get("type") or "")[:20]
            vstatus = str(item.get("status") or "")[:20]
            currency = str(item.get("currency") or "")[:10]

            cursor.execute(
                "SELECT id FROM vendor WHERE vendor_code=%s",
                (code,),
            )
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    """UPDATE vendor SET
                        vendor_name=%s, vendor_type=%s, status=%s,
                        currency=%s, sync_time=%s, del_flag=0, update_time=%s
                       WHERE id=%s""",
                    (name, vtype, vstatus, currency, now, now, existing["id"]),
                )
                stats["updated"] += 1
            else:
                cursor.execute(
                    """INSERT INTO vendor
                        (id, vendor_code, vendor_name, vendor_type,
                         status, currency, sync_time, create_time, del_flag)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,0)""",
                    (generate_id(), code, name, vtype, vstatus, currency, now, now),
                )
                stats["inserted"] += 1

        conn.commit()
        print(f"[OK] 供应商同步完成: {stats}")
        return stats

    except Exception as e:
        conn.rollback()
        print(f"[ERROR] 供应商同步失败: {e}")
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        conn.close()


# ── Excel 导出 ────────────────────────────────────────────────────────────────


def export_vendors_excel(
    keyword: Optional[str] = None,
) -> bytes:
    """
    将 vendor 表数据导出为 Excel（字节流），供客户导入使用

    Args:
        keyword: 供应商编号或名称关键字过滤

    Returns:
        Excel 文件字节内容
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        where = ["del_flag=0"]
        params = []
        if keyword:
            where.append("(vendor_code LIKE %s OR vendor_name LIKE %s)")
            params += [f"%{keyword}%", f"%{keyword}%"]

        cursor.execute(
            f"""SELECT vendor_code, vendor_name, vendor_type, status, currency, sync_time
                FROM vendor WHERE {' AND '.join(where)}
                ORDER BY vendor_code""",
            params,
        )
        rows = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应商账户"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["供应商编号", "供应商名称", "供应商类型", "状态", "币种", "同步时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append([
            r["vendor_code"],
            r["vendor_name"] or "",
            r["vendor_type"] or "",
            r["status"] or "",
            r["currency"] or "",
            str(r["sync_time"])[:19] if r["sync_time"] else "",
        ])

    for col, w in zip("ABCDEF", [20, 40, 15, 12, 10, 20]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
