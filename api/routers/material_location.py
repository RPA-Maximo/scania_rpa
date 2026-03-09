"""
物料默认仓库仓位关联表 API 路由

功能：
  - Excel 文件导入（物料编号 + 默认货柜，仓库自动从 bin_inventory/warehouse_bin 推导）
  - 从 Maximo MXAPIINVENTORY 的 defaultbin 字段同步（优先级低于 Excel 手工导入）
  - 列表分页查询
  - 单条更新/删除
  - 下载 Excel 导入模板
  - 导出当前数据为 Excel

Excel 格式（导入）：
  列 A: 物料编号   (必填)
  列 B: 货柜编号   (必填)
  列 C: 备注       (可选)

端点：
  POST   /api/material-location/import    上传 Excel 导入（优先级最高）
  POST   /api/material-location/sync      从 Maximo 同步缺省货柜（不覆盖 Excel 导入数据）
  GET    /api/material-location           分页查询列表
  PUT    /api/material-location/{id}      修改单条（货柜/备注）
  DELETE /api/material-location/{id}      软删除
  GET    /api/material-location/template  下载导入模板
  GET    /api/material-location/export    导出当前数据
"""
import io
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import openpyxl
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from src.utils.db import get_connection, generate_id
from src.sync.material_location_sync import sync_material_locations

router = APIRouter(prefix="/api/material-location", tags=["material-location"])


# ── 列名别名映射（兼容中英文列头）────────────────────────────────────────────

_ITEM_ALIASES = {"物料编号", "item_number", "itemnum", "物料", "item"}
_BIN_ALIASES  = {"货柜编号", "bin_code", "binnum", "货柜", "缺省货柜", "default_bin", "defaultbin"}
_REMARK_ALIASES = {"备注", "remark", "remarks", "note", "notes"}


def _match_col(header: str, aliases: set) -> bool:
    return header.strip().lower() in {a.lower() for a in aliases}


# ── 辅助：按 bin_code 推导仓库 ───────────────────────────────────────────────

def _derive_warehouse(cursor, bin_code: str) -> Optional[str]:
    """
    按货柜编号推导仓库编码，查询顺序：
      1. bin_inventory（含实时库存的货柜）
      2. warehouse_bin （所有已同步的仓位主数据，无库存时仍有记录）
    """
    if not bin_code:
        return None
    cursor.execute(
        """SELECT warehouse FROM bin_inventory
           WHERE bin_code=%s AND del_flag=0
             AND warehouse IS NOT NULL AND warehouse!=''
           LIMIT 1""",
        (bin_code,),
    )
    row = cursor.fetchone()
    if row:
        return row["warehouse"]
    # 回查仓位主数据表（Maximo 仓位同步，无论有无库存均有记录）
    cursor.execute(
        """SELECT warehouse FROM warehouse_bin
           WHERE bin_code=%s AND del_flag=0
           LIMIT 1""",
        (bin_code,),
    )
    row = cursor.fetchone()
    return row["warehouse"] if row else None


def _derive_bin_name(cursor, bin_code: str) -> Optional[str]:
    """从 warehouse_bin 查货柜名称，无则回查 bin_inventory"""
    if not bin_code:
        return None
    cursor.execute(
        "SELECT bin_name FROM warehouse_bin WHERE bin_code=%s AND del_flag=0 LIMIT 1",
        (bin_code,),
    )
    row = cursor.fetchone()
    if row and row["bin_name"]:
        return row["bin_name"]
    cursor.execute(
        "SELECT bin_name FROM bin_inventory WHERE bin_code=%s AND del_flag=0 LIMIT 1",
        (bin_code,),
    )
    row = cursor.fetchone()
    return row["bin_name"] if row else None


def _derive_item_name(cursor, item_number: str) -> Optional[str]:
    """从 material 表查物料名称"""
    if not item_number:
        return None
    cursor.execute(
        "SELECT name FROM material WHERE code=%s AND del_flag=0 LIMIT 1",
        (item_number,),
    )
    row = cursor.fetchone()
    return row["name"] if row else None


# ── 端点 ──────────────────────────────────────────────────────────────────────


@router.post("/import", summary="Excel 导入物料默认仓位")
async def import_from_excel(file: UploadFile = File(...)):
    """
    上传 Excel 文件批量导入物料默认货柜关联。

    **Excel 格式：**
    - 第1行为列标题（物料编号、货柜编号、备注）
    - 第2行起为数据

    **业务规则：**
    - 以 `item_number`（物料编号）为唯一键，已存在则更新，不存在则插入
    - 仓库自动从 `bin_inventory` 表中按货柜编号推导（无匹配则留空，记录警告）
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel 文件解析失败: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 无数据行（至少需要标题行+1行数据）")

    # 解析列索引
    headers = [str(c or "").strip() for c in rows[0]]
    item_idx = bin_idx = remark_idx = None
    for i, h in enumerate(headers):
        if _match_col(h, _ITEM_ALIASES):
            item_idx = i
        elif _match_col(h, _BIN_ALIASES):
            bin_idx = i
        elif _match_col(h, _REMARK_ALIASES):
            remark_idx = i

    if item_idx is None:
        raise HTTPException(status_code=400, detail="未找到"物料编号"列（支持列名：物料编号、item_number、itemnum）")
    if bin_idx is None:
        raise HTTPException(status_code=400, detail="未找到"货柜编号"列（支持列名：货柜编号、bin_code、binnum、缺省货柜）")

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    stats = {"inserted": 0, "updated": 0, "skipped": 0, "warnings": []}
    now = datetime.now()

    try:
        for row_idx, row in enumerate(rows[1:], start=2):
            item_number = str(row[item_idx] or "").strip()
            bin_code    = str(row[bin_idx]  or "").strip()

            if not item_number or not bin_code:
                stats["skipped"] += 1
                continue

            remark = str(row[remark_idx] or "").strip() if remark_idx is not None else ""

            # 推导仓库、货柜名称、物料名称
            warehouse = _derive_warehouse(cursor, bin_code)
            if not warehouse:
                stats["warnings"].append(
                    f"第{row_idx}行 物料{item_number}: 货柜{bin_code}在bin_inventory中未找到，仓库留空"
                )
            bin_name  = _derive_bin_name(cursor, bin_code)
            item_name = _derive_item_name(cursor, item_number)

            # Upsert
            cursor.execute(
                "SELECT id FROM material_location WHERE item_number=%s",
                (item_number,),
            )
            existing = cursor.fetchone()

            if existing:
                cursor.execute(
                    """UPDATE material_location SET
                        item_name=%s, warehouse=%s, bin_code=%s, bin_name=%s,
                        remark=%s, import_time=%s, import_source='excel',
                        update_time=%s, del_flag=0
                       WHERE id=%s""",
                    (item_name, warehouse, bin_code, bin_name, remark, now, now, existing["id"]),
                )
                stats["updated"] += 1
            else:
                cursor.execute(
                    """INSERT INTO material_location
                        (id, item_number, item_name, warehouse, bin_code, bin_name,
                         remark, import_time, import_source, create_time, del_flag)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'excel',%s,0)""",
                    (generate_id(), item_number, item_name, warehouse, bin_code,
                     bin_name, remark, now, now),
                )
                stats["inserted"] += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败: {e}")
    finally:
        cursor.close()
        conn.close()

    return {
        "success":  True,
        "inserted": stats["inserted"],
        "updated":  stats["updated"],
        "skipped":  stats["skipped"],
        "warnings": stats["warnings"],
        "message":  (
            f"导入完成：新增 {stats['inserted']} 条，"
            f"更新 {stats['updated']} 条，"
            f"跳过 {stats['skipped']} 条"
        ),
    }


class SyncRequest(BaseModel):
    warehouse:  Optional[str] = None
    site_id:    Optional[str] = None
    max_pages:  int = 50
    page_size:  int = 100


@router.post("/sync", summary="从 Maximo 同步物料缺省货柜")
def sync_from_maximo(req: SyncRequest):
    """
    从 Maximo MXAPIINVENTORY 的 **缺省货柜（defaultbin）** 字段自动同步到
    `material_location` 表。

    **优先级规则：**
    - 已通过 Excel 手工导入的记录（`import_source='excel'`）不会被本接口覆盖
    - 仅写入 `import_source='maximo'` 的记录或全新记录

    **仓库推导逻辑（同 Excel 导入）：**
    1. 查 `bin_inventory`（有实时库存的货柜）
    2. 回查 `warehouse_bin`（Maximo 仓位主数据，无库存时仍有记录）
    """
    try:
        stats = sync_material_locations(
            warehouse=req.warehouse,
            site_id=req.site_id,
            max_pages=req.max_pages,
            page_size=req.page_size,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Maximo 同步失败: {e}")

    return {
        "success":      True,
        "inserted":     stats["inserted"],
        "updated":      stats["updated"],
        "skipped":      stats["skipped"],
        "no_warehouse": stats["no_warehouse"],
        "message": (
            f"同步完成：新增 {stats['inserted']} 条，"
            f"更新 {stats['updated']} 条，"
            f"跳过 {stats['skipped']} 条"
            + (f"，{stats['no_warehouse']} 条货柜未找到仓库" if stats["no_warehouse"] else "")
        ),
    }


@router.get("", summary="查询物料默认仓位列表")
def list_material_locations(
    keyword:   Optional[str] = Query(None, description="物料编号或名称关键字"),
    warehouse: Optional[str] = Query(None, description="仓库过滤"),
    page:      int           = Query(1,    ge=1),
    page_size: int           = Query(20,   ge=1, le=200),
):
    """分页查询物料默认仓库仓位关联列表"""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        where  = ["del_flag=0"]
        params = []

        if keyword:
            where.append("(item_number LIKE %s OR item_name LIKE %s)")
            params += [f"%{keyword}%", f"%{keyword}%"]
        if warehouse:
            where.append("warehouse=%s")
            params.append(warehouse)

        w_sql  = " AND ".join(where)
        offset = (page - 1) * page_size

        cursor.execute(f"SELECT COUNT(*) AS total FROM material_location WHERE {w_sql}", params)
        total = cursor.fetchone()["total"]

        cursor.execute(
            f"""SELECT id, item_number, item_name, warehouse, bin_code, bin_name,
                       remark, import_time, update_time
                FROM material_location
                WHERE {w_sql}
                ORDER BY item_number
                LIMIT %s OFFSET %s""",
            params + [page_size, offset],
        )
        rows = cursor.fetchall()

        for r in rows:
            for f in ("import_time", "update_time"):
                if r.get(f) and not isinstance(r[f], str):
                    r[f] = r[f].strftime("%Y-%m-%d %H:%M:%S")

        return {"total": total, "page": page, "page_size": page_size, "items": rows}
    finally:
        cursor.close()
        conn.close()


class LocationUpdateRequest(BaseModel):
    bin_code:  Optional[str] = None
    remark:    Optional[str] = None


@router.put("/{loc_id}", summary="修改物料默认货柜")
def update_location(loc_id: int, req: LocationUpdateRequest):
    """
    修改指定记录的默认货柜编号（仓库自动重新推导）
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT id FROM material_location WHERE id=%s AND del_flag=0",
            (loc_id,),
        )
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="记录不存在")

        fields, params = [], []
        if req.bin_code is not None:
            bin_code  = req.bin_code.strip()
            warehouse = _derive_warehouse(cursor, bin_code)
            bin_name  = _derive_bin_name(cursor, bin_code)
            fields += ["bin_code=%s", "warehouse=%s", "bin_name=%s"]
            params += [bin_code, warehouse, bin_name]

        if req.remark is not None:
            fields.append("remark=%s")
            params.append(req.remark)

        if not fields:
            raise HTTPException(status_code=400, detail="未提供任何修改字段")

        fields.append("update_time=NOW()")
        params.append(loc_id)
        cursor.execute(
            f"UPDATE material_location SET {', '.join(fields)} WHERE id=%s",
            params,
        )
        conn.commit()
        return {"success": True, "message": "更新成功"}
    finally:
        cursor.close()
        conn.close()


@router.delete("/{loc_id}", summary="删除物料默认仓位记录")
def delete_location(loc_id: int):
    """软删除（del_flag=1）"""
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE material_location SET del_flag=1, update_time=NOW() WHERE id=%s",
            (loc_id,),
        )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="记录不存在")
        conn.commit()
        return {"success": True}
    finally:
        cursor.close()
        conn.close()


@router.get("/template", summary="下载导入模板")
def download_template():
    """下载 Excel 导入模板（含列标题和示例数据）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物料默认仓位导入"

    # 标题行
    ws.append(["物料编号", "货柜编号", "备注"])

    # 示例行
    ws.append(["20326796", "CVC4802A", "示例数据，请替换"])
    ws.append(["20326797", "CVC4803B", ""])

    # 列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=material_location_template.xlsx"},
    )


@router.get("/export", summary="导出当前数据为 Excel")
def export_locations(
    warehouse: Optional[str] = Query(None, description="仓库过滤"),
    keyword:   Optional[str] = Query(None, description="物料编号或名称关键字"),
):
    """将当前 material_location 数据导出为 Excel"""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        where  = ["del_flag=0"]
        params = []
        if keyword:
            where.append("(item_number LIKE %s OR item_name LIKE %s)")
            params += [f"%{keyword}%", f"%{keyword}%"]
        if warehouse:
            where.append("warehouse=%s")
            params.append(warehouse)

        cursor.execute(
            f"""SELECT item_number, item_name, warehouse, bin_code, bin_name,
                       remark, import_time
                FROM material_location WHERE {' AND '.join(where)}
                ORDER BY item_number""",
            params,
        )
        rows = cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "物料默认仓位"
    ws.append(["物料编号", "物料名称", "默认仓库", "默认货柜", "货柜名称", "备注", "导入时间"])
    for r in rows:
        ws.append([
            r["item_number"],
            r["item_name"] or "",
            r["warehouse"] or "",
            r["bin_code"],
            r["bin_name"] or "",
            r["remark"] or "",
            str(r["import_time"])[:19] if r["import_time"] else "",
        ])

    for col, w in zip("ABCDEFG", [20, 30, 15, 15, 20, 30, 20]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=material_location_{ts}.xlsx"},
    )
