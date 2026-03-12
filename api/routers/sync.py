"""
PO 增量同步管理 API 路由

提供对自动同步调度器的完整控制：状态查询、手动触发、参数调整。
"""
import io
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, validator
from typing import List

from src.sync.po_sync_service import po_sync_service, po_sync_scheduler
from src.utils.db import get_connection

router = APIRouter(prefix="/api/sync/po", tags=["PO 增量同步"])


# ── 请求模型 ──────────────────────────────────────────────────────────────────

class SyncConfigRequest(BaseModel):
    status_filter: Optional[str] = Field(
        None,
        description="PO 状态筛选，如 'APPR'（已批准）、'WAPPR'（待批准）。"
                    "留空则不按状态过滤",
    )
    max_pages: Optional[int] = Field(
        None,
        description="每次同步最多抓取的页数（默认 5 页 × 每页 20 条 = 最多 100 条）",
        ge=1, le=100,
    )
    page_size: Optional[int] = Field(
        None,
        description="每页条数（默认 20）",
        ge=1, le=100,
    )
    auto_sync_materials: Optional[bool] = Field(
        None,
        description="是否自动将 Maximo 新物料同步到 WMS material 表（默认 true）",
    )


class IntervalRequest(BaseModel):
    interval_minutes: float = Field(
        ...,
        description="同步间隔（分钟）。最小 1 分钟，默认 5 分钟",
        ge=1, le=1440,
    )


# ── 端点 ──────────────────────────────────────────────────────────────────────

@router.get("/status", summary="查询同步状态")
async def get_sync_status():
    """
    返回同步服务和调度器的完整状态，包括：
    - 调度器是否运行、当前同步间隔
    - 上次同步时间、结果摘要
    - 当前同步参数配置
    """
    return {
        'scheduler': po_sync_scheduler.get_status(),
        'service': po_sync_service.get_status(),
    }


@router.post("/trigger", summary="手动触发一次同步")
async def trigger_sync():
    """
    立即执行一次增量同步（不影响自动调度计划）。

    - 若上次同步仍在运行，本次将被跳过并返回提示
    - 认证信息从 auth_manager 获取，需提前通过 `POST /api/auth/curl` 更新
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        result = await loop.run_in_executor(executor, po_sync_service.sync_once)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"同步执行异常: {e}")
    finally:
        executor.shutdown(wait=False)

    if not result.get('success') and not result.get('skipped'):
        raise HTTPException(status_code=500, detail=result.get('message', '同步失败'))
    return result


@router.post("/start", summary="启动自动同步调度器")
async def start_scheduler():
    """启动 5 分钟定时自动同步调度器（服务启动时已自动调用，通常无需手动操作）"""
    if po_sync_scheduler.get_status()['running']:
        return {'message': '调度器已在运行中', 'status': po_sync_scheduler.get_status()}
    po_sync_scheduler.start()
    return {'message': '调度器已启动', 'status': po_sync_scheduler.get_status()}


@router.post("/stop", summary="停止自动同步调度器")
async def stop_scheduler():
    """停止定时自动同步（不影响正在执行的同步任务）"""
    if not po_sync_scheduler.get_status()['running']:
        return {'message': '调度器未在运行', 'status': po_sync_scheduler.get_status()}
    po_sync_scheduler.stop()
    return {'message': '调度器已停止', 'status': po_sync_scheduler.get_status()}


@router.put("/config", summary="更新同步参数")
async def update_sync_config(request: SyncConfigRequest):
    """
    动态调整同步参数，立即生效（下次同步触发时使用新参数）。

    **常用配置示例：**

    只同步已批准的 PO，每次最多 200 条：
    ```json
    {"status_filter": "APPR", "max_pages": 10, "page_size": 20}
    ```

    不限状态全量扫描：
    ```json
    {"status_filter": null, "max_pages": 50}
    ```
    """
    result = po_sync_service.update_config(
        status_filter=request.status_filter,
        max_pages=request.max_pages,
        page_size=request.page_size,
        auto_sync_materials=request.auto_sync_materials,
    )
    return result


@router.put("/interval", summary="修改同步间隔")
async def update_sync_interval(request: IntervalRequest):
    """
    修改自动同步的时间间隔（分钟），立即对下次触发生效。

    默认 5 分钟。建议范围：1～60 分钟。
    """
    seconds = int(request.interval_minutes * 60)
    po_sync_scheduler.set_interval(seconds)
    return {
        'message': f'同步间隔已修改为 {request.interval_minutes} 分钟',
        'interval_seconds': seconds,
        'scheduler_status': po_sync_scheduler.get_status(),
    }


@router.get("/export", summary="导出采购订单为 Excel")
async def export_po_excel(
    limit: int = Query(1000, ge=1, le=10000, description="最多导出条数（默认 1000，最大 10000）"),
):
    """
    将数据库中已同步的采购订单导出为 Excel 文件，包含两个工作表：

    - **采购订单**：主表信息（PO 号、供应商、收货方、创建时间等）
    - **采购明细**：子表行项目（物料号、描述、数量、单价等）

    点击 Swagger UI 的 **Download file** 按钮即可保存文件。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，请执行: pip install openpyxl")

    try:
        conn = get_connection()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"数据库连接失败: {e}")

    try:
        cursor = conn.cursor(dictionary=True)

        # ── 查询主表 ──────────────────────────────────────────────────────
        cursor.execute(f"""
            SELECT
                id,
                code            AS `PO号`,
                description     AS `描述`,
                status          AS `状态`,
                order_date      AS `订单日期`,
                request_date    AS `需求日期`,
                total_cost      AS `总金额`,
                currency        AS `币种`,
                supplier_name   AS `供应商名称`,
                vendor_code     AS `供应商代码`,
                supplier_address AS `供应商地址`,
                supplier_zip    AS `供应商邮政编码`,
                supplier_city   AS `供应商城市`,
                supplier_contact AS `供应商联系人`,
                supplier_phone  AS `供应商电话`,
                supplier_email  AS `供应商邮箱`,
                company_name    AS `公司名称`,
                street_address  AS `街道地址`,
                city            AS `城市`,
                postal_code     AS `邮政编码`,
                country         AS `国家`,
                create_time     AS `同步时间`
            FROM purchase_order
            WHERE del_flag = 0
            ORDER BY create_time DESC
            LIMIT {limit}
        """)
        headers = cursor.fetchall()

        # ── 查询子表 ──────────────────────────────────────────────────────
        if headers:
            po_ids = [r['id'] for r in headers]
            placeholders = ','.join(['%s'] * len(po_ids))
            cursor.execute(f"""
                SELECT
                    p.code              AS `PO号`,
                    b.number            AS `行号`,
                    COALESCE(b.item_code, m.code) AS `物料编号`,
                    b.sku_names         AS `物料名称`,
                    b.model_num         AS `型号`,
                    b.size_info         AS `尺寸`,
                    b.qty               AS `数量`,
                    b.ordering_unit     AS `订购单位`,
                    b.target_container  AS `目标货柜`,
                    w.code              AS `目标仓库`,
                    b.form_id           AS `主表ID`
                FROM purchase_order_bd b
                JOIN purchase_order p ON p.id = b.form_id
                LEFT JOIN material m ON m.id = b.sku
                LEFT JOIN warehouse w ON w.id = b.warehouse
                WHERE b.form_id IN ({placeholders})
                ORDER BY b.form_id, b.number
            """, po_ids)
            details = cursor.fetchall()
        else:
            details = []

        cursor.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"查询数据失败: {e}")
    finally:
        conn.close()

    # ── 生成 Excel ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_sheet(ws, rows: list):
        if not rows:
            ws.append(["暂无数据"])
            return
        # 写表头（去掉内部 id/del_flag 字段）
        cols = [k for k in rows[0].keys() if k not in ('id', 'del_flag', '主表ID')]
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        ws.row_dimensions[1].height = 22
        # 写数据
        for row in rows:
            ws.append([row.get(c) for c in cols])
        # 自适应列宽（最大 40）
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    ws1 = wb.active
    ws1.title = "采购订单"
    _write_sheet(ws1, headers)

    ws2 = wb.create_sheet("采购明细")
    _write_sheet(ws2, details)

    # ── 输出为字节流 ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"PO_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


class ResyncRequest(BaseModel):
    po_numbers: List[str] = Field(
        ...,
        description="要强制重同步的 PO 号列表，如 ['CN4300', 'CN5044']",
        min_items=1,
    )

    @validator('po_numbers', each_item=True)
    def validate_po_number(cls, v):
        v = v.strip()
        if not v or v.lower() == 'string' or len(v) < 3:
            raise ValueError(f"无效的 PO 号: '{v}'，请填写真实的 PO 号（如 CN4300）")
        return v


@router.post("/resync", summary="强制重新同步指定PO（修复历史缺失字段）")
async def resync_specific_pos(request: ResyncRequest):
    """
    对指定 PO 号执行**强制重同步**：先删除 WMS 中已有记录，再从 Maximo 重新拉取写入。

    **适用场景：**
    - 字段结构变更后（新增 item_code / model_num / size_info / target_container），
      修复在变更前已同步的历史数据
    - 某个 PO 数据有误，需要覆盖更新

    **注意：** 同步期间该 PO 的数据会短暂不可见（删后再插）。
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        result = await loop.run_in_executor(
            executor, lambda: po_sync_service.resync_pos(request.po_numbers)
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"重同步执行异常: {e}")
    finally:
        executor.shutdown(wait=False)

    if not result.get('success') and not result.get('skipped'):
        fetch_failed = result.get('fetch_failed', [])
        if fetch_failed and len(fetch_failed) == len(request.po_numbers):
            raise HTTPException(
                status_code=422,
                detail=f"指定的 PO 号在 Maximo 中未找到: {fetch_failed}",
            )
        raise HTTPException(status_code=500, detail=result.get('message', '重同步失败'))
    return result


@router.post("/resync-all", summary="强制重新同步数据库中所有PO（批量修复）")
async def resync_all_pos():
    """
    将数据库中**所有** PO 从 Maximo 重新拉取并覆盖写入。

    **适用场景：** 字段结构变更后，批量修复所有历史记录中的缺失字段。

    **警告：** 如果 PO 数量较多（数百条以上），此操作耗时较长，请耐心等待。
    建议先用 `/resync` 针对少量 PO 测试验证，确认正常后再执行全量重同步。

    请求超时设置为 10 分钟。
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        result = await loop.run_in_executor(
            executor, po_sync_service.resync_all_existing
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"全量重同步异常: {e}")
    finally:
        executor.shutdown(wait=False)

    if not result.get('success') and not result.get('skipped'):
        raise HTTPException(status_code=500, detail=result.get('message', '重同步失败'))
    return result


@router.get("/debug/poline-raw/{po_number}", summary="诊断：用poline{*}拉取全部字段（含cx自定义字段）")
async def debug_poline_raw_fields(po_number: str):
    """
    用 `poline{*}` 查询指定 PO 的原始明细，返回第一行的**全部字段**（含 Scania 自定义 cx 字段）。

    **解决的问题：** 当前 po_fetcher 明确列出了 poline 字段，可能漏掉了 Scania 自定义字段。
    本接口绕过 po_fetcher，直接用 `poline{*}` 查询，找出型号/尺寸对应的真实字段名。

    **对比 /debug/poline/{po_number}（用 po_fetcher 的固定 select）**：
    - 本接口: `poline{*}` → 显示所有字段（目的：找缺失的 cx 字段名）
    - 那个接口: `poline{...固定列表...}` → 显示当前已配置字段的值
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    from config import get_maximo_auth, DEFAULT_HEADERS
    from config.settings import MAXIMO_BASE_URL, VERIFY_SSL
    from config.settings_manager import settings_manager
    import requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    try:
        auth = get_maximo_auth()
    except ValueError as e:
        raise HTTPException(status_code=401, detail=f"认证未配置: {e}")

    headers = {
        **DEFAULT_HEADERS,
        'Cookie': auth['cookie'],
        'x-csrf-token': auth['csrf_token'],
    }

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    def _fetch():
        url = f"{MAXIMO_BASE_URL}/oslc/os/MXAPIPO"
        params = {
            'oslc.select': '*,poline{*}',
            'oslc.where': f'ponum="{po_number}"',
            '_dropnulls': '0',
            'oslc.pageSize': 1,
        }
        return requests.get(
            url, headers=headers, params=params,
            verify=VERIFY_SSL,
            proxies=settings_manager.get_proxies(),
            timeout=120,
        )

    try:
        resp = await loop.run_in_executor(executor, _fetch)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"请求失败: {e}")
    finally:
        executor.shutdown(wait=False)

    if resp.status_code != 200:
        raise HTTPException(
            status_code=resp.status_code,
            detail=f"Maximo 返回 {resp.status_code}: {resp.text[:300]}",
        )

    data = resp.json()
    items = data.get('member') or data.get('rdfs:member') or []
    if not items:
        raise HTTPException(status_code=404, detail=f"未找到 PO: {po_number}")

    # 标准化（去命名空间前缀）
    def _normalize(obj):
        if isinstance(obj, dict):
            return {(k.split(':', 1)[1] if ':' in k else k): _normalize(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [_normalize(i) for i in obj]
        return obj

    po = _normalize(items[0])
    poline = po.get('poline', [])

    if not poline:
        return {'po_number': po_number, 'poline_count': 0, 'message': '无明细行'}

    # 取第一条明细，列出全部字段
    first_line = poline[0]
    all_keys = sorted(first_line.keys())
    fields_with_values = {k: v for k, v in first_line.items() if v is not None and v != ''}
    null_fields = [k for k, v in first_line.items() if v is None or v == '']

    # 所有行里有值的字段（聚合）
    all_line_fields_with_value: set = set()
    for line in poline:
        for k, v in line.items():
            if v is not None and v != '':
                all_line_fields_with_value.add(k)

    # 找第一行有 cx 前缀的字段
    cx_fields_in_first_line = {k: v for k, v in first_line.items() if k.lower().startswith('cx')}
    cx_fields_with_value_any_line = sorted(
        k for k in all_line_fields_with_value if k.lower().startswith('cx')
    )

    return {
        'po_number': po_number,
        'poline_count': len(poline),
        'first_line_all_keys': all_keys,
        'first_line_fields_with_values': fields_with_values,
        'first_line_null_fields': sorted(null_fields),
        'cx_fields_in_first_line': cx_fields_in_first_line,
        'cx_fields_with_value_any_line': cx_fields_with_value_any_line,
        'all_fields_with_value_any_line': sorted(all_line_fields_with_value),
        'hint': (
            '重点看 cx_fields_with_value_any_line —— 这些是所有 poline 行中有值的 Scania 自定义字段，'
            '其中很可能包含型号/尺寸对应的真实字段名。'
        ),
    }


@router.get("/debug/poline/{po_number}", summary="诊断：查看指定PO的原始poline字段")
async def debug_poline_fields(po_number: str):
    """
    从 Maximo 实时抓取指定采购订单，扫描**全部** poline 行，
    汇报哪些字段有实际值（区分"字段缺失"和"字段存在但值为空"）。

    **使用方法：**
    - 填入一个已知有尺寸/型号数据的 PO 号（如 `CN5074`）
    - 查看 `fields_with_values` — 在所有行中出现过非空值的字段
    - 若 `newitemdesc` 不在 `fields_with_values` 里，说明该字段在 Maximo 里就是空的
    - 若 `newitemdesc` 完全不在 `all_fields_seen` 里，说明 OSLC 资源未配置该字段
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    from src.fetcher.po_fetcher import fetch_po_by_number

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    try:
        po_data = await loop.run_in_executor(
            executor, lambda: fetch_po_by_number(po_number, save_to_file=False)
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"抓取 Maximo 数据失败: {e}")
    finally:
        executor.shutdown(wait=False)

    if not po_data:
        raise HTTPException(status_code=404, detail=f"未找到 PO: {po_number}（可能认证过期或不存在）")

    poline = po_data.get('poline', [])
    if not poline:
        return {'po_number': po_number, 'poline_count': 0, 'message': '该PO无明细行'}

    # 扫描全部行，统计字段出现情况
    all_fields_seen: set = set()      # 在任意行出现的字段名（含 null）
    fields_with_values: set = set()   # 在任意行有非空值的字段名

    for line in poline:
        all_fields_seen.update(line.keys())
        for k, v in line.items():
            if v is not None and v != '':
                fields_with_values.add(k)

    # 找3条有 newitemdesc 或 catalogcode 值的样本行
    target_fields = ('newitemdesc', 'catalogcode')
    sample_lines = []
    for line in poline:
        if any(line.get(f) for f in target_fields):
            sample_lines.append({
                'itemnum':     line.get('itemnum'),
                'description': line.get('description'),
                'catalogcode': line.get('catalogcode'),
                'newitemdesc': line.get('newitemdesc'),
                'location':    line.get('location'),
                'storeloc':    line.get('storeloc'),
                'linetype':    line.get('linetype'),
            })
        if len(sample_lines) >= 3:
            break

    # 如果没有找到有值的行，取前3行作为样本
    if not sample_lines:
        for line in poline[:3]:
            sample_lines.append({
                'all_keys':    sorted(line.keys()),
                'itemnum':     line.get('itemnum'),
                'description': line.get('description'),
                'catalogcode': line.get('catalogcode'),
                'newitemdesc': line.get('newitemdesc'),
                'location':    line.get('location'),
                'storeloc':    line.get('storeloc'),
                'linetype':    line.get('linetype'),
            })

    KEY_FIELDS = ['itemnum', 'catalogcode', 'newitemdesc', 'location',
                  'storeloc', 'orderqty', 'orderunit', 'description']
    return {
        'po_number':        po_number,
        'poline_count':     len(poline),
        'all_fields_seen':  sorted(all_fields_seen),
        'fields_with_values': sorted(fields_with_values),
        'key_field_status': {
            f: {
                'in_oslc':   f in all_fields_seen,
                'has_value': f in fields_with_values,
            }
            for f in KEY_FIELDS
        },
        'sample_lines_with_data': sample_lines,
    }


@router.get("/debug/scan-fields", summary="诊断：扫描多个PO，找有指定字段值的样本")
async def scan_po_fields(
    fields: str = Query(
        "catalogcode,newitemdesc",
        description="要检查的字段名，逗号分隔，如 'catalogcode,newitemdesc,location'",
    ),
    max_scan: int = Query(
        30,
        description="最多扫描的PO数量（从数据库取最近的PO逐个检查）",
        ge=1, le=200,
    ),
    max_hits: int = Query(
        5,
        description="找到多少个命中PO后停止",
        ge=1, le=20,
    ),
):
    """
    从数据库取最近的 PO，逐个从 Maximo 拉取 poline，
    找出哪些 PO 在指定字段上有实际值。

    **用途：** 确认 `catalogcode` / `newitemdesc` / `location` 字段
    在真实数据里是否存在，以及在哪些 PO 上。
    """
    import asyncio
    import time
    from concurrent.futures import ThreadPoolExecutor
    from src.fetcher.po_fetcher import fetch_po_by_number

    target_fields = [f.strip() for f in fields.split(',') if f.strip()]
    if not target_fields:
        raise HTTPException(status_code=400, detail="fields 参数不能为空")

    # 从 DB 取最近的 PO 编号
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT code FROM purchase_order ORDER BY create_time DESC LIMIT %s",
            (max_scan,),
        )
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        po_codes = [r[0] for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"查询数据库失败: {e}")

    if not po_codes:
        return {'message': '数据库中无PO记录', 'hits': []}

    hits = []
    scanned = 0
    # 保留第一个成功获取到的 PO 概况（用于验证扫描在正常工作）
    first_po_probe = None
    executor = ThreadPoolExecutor(max_workers=1)
    loop = asyncio.get_event_loop()

    for po_code in po_codes:
        if len(hits) >= max_hits:
            break
        scanned += 1
        try:
            po_data = await loop.run_in_executor(
                executor, lambda c=po_code: fetch_po_by_number(c, save_to_file=False)
            )
        except Exception:
            continue

        if not po_data:
            continue

        poline = po_data.get('poline', [])

        # 记录第一个成功 PO 的概况
        if first_po_probe is None and poline:
            sample_line = poline[0]
            first_po_probe = {
                'po_number':   po_code,
                'poline_count': len(poline),
                'first_line_keys': sorted(sample_line.keys()),
                'first_line_target_values': {
                    f: sample_line.get(f) for f in target_fields
                },
            }

        for line in poline:
            matched = {f: line.get(f) for f in target_fields if line.get(f)}
            if matched:
                hits.append({
                    'po_number':  po_code,
                    'polinenum':  line.get('polinenum'),
                    'itemnum':    line.get('itemnum'),
                    'description': line.get('description'),
                    'linetype':   line.get('linetype'),
                    'matched_fields': matched,
                })
                break  # 每个 PO 只取第一个命中行

        time.sleep(0.3)

    executor.shutdown(wait=False)

    return {
        'target_fields': target_fields,
        'scanned_po_count': scanned,
        'hit_po_count': len(hits),
        'hits': hits,
        'note': '已扫描完所有PO未找到目标字段' if not hits else '',
        # 当0命中时，用此字段确认扫描确实在工作（字段是否出现在 OSLC 响应里）
        'scan_probe_first_po': first_po_probe,
    }


@router.get("/debug/mxitem/{item_num}", summary="诊断：查询MXITEM/MXAPIITEM物料主数据字段")
async def debug_mxitem_fields(item_num: str):
    """
    通过 OSLC 查询指定物料的完整字段，
    找出型号（catalogcode / manufacturernum / modelnum）、规格（newitemdesc）
    等字段是否存在于物料主数据里。

    依次尝试对象名：`MXAPIITEM` → `MXITEM` → `MXAPIITMSPEC`，
    返回第一个成功的结果及所有尝试的状态码（便于排查 400）。

    **用途：** 当 MXAPIPO poline 中 catalogcode/newitemdesc 为空时，
    看物料主数据是否有对应的型号/规格信息。
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    from config import get_maximo_auth, DEFAULT_HEADERS
    from config.settings import MAXIMO_BASE_URL, VERIFY_SSL
    from config.settings_manager import settings_manager
    import requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    try:
        auth = get_maximo_auth()
    except ValueError as e:
        raise HTTPException(status_code=401, detail=f"认证未配置: {e}")

    headers = {
        **DEFAULT_HEADERS,
        'Cookie': auth['cookie'],
        'x-csrf-token': auth['csrf_token'],
    }

    # Maximo 不同版本/配置下物料对象名不同，依次尝试
    CANDIDATE_OBJECTS = ['MXAPIITEM', 'MXITEM', 'mxapiitem']

    attempt_log = []

    def _try_fetch(obj_name: str):
        url = f"{MAXIMO_BASE_URL}/oslc/os/{obj_name}"
        params = {
            'oslc.select': '*',
            'oslc.where': f'itemnum="{item_num}"',
            '_dropnulls': '0',
        }
        resp = requests.get(
            url, headers=headers, params=params,
            verify=VERIFY_SSL,
            proxies=settings_manager.get_proxies(),
            timeout=60,
        )
        return resp

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)
    success_resp = None
    used_object = None

    for obj_name in CANDIDATE_OBJECTS:
        try:
            resp = await loop.run_in_executor(
                executor, lambda o=obj_name: _try_fetch(o)
            )
        except Exception as e:
            attempt_log.append({'object': obj_name, 'status': 'exception', 'error': str(e)})
            continue

        entry = {
            'object': obj_name,
            'status_code': resp.status_code,
        }
        if resp.status_code == 400:
            try:
                entry['error_body'] = resp.json()
            except Exception:
                entry['error_body'] = resp.text[:500]
        attempt_log.append(entry)

        if resp.status_code == 200:
            success_resp = resp
            used_object = obj_name
            break
        if resp.status_code == 401:
            break  # 认证问题，不必继续尝试

    executor.shutdown(wait=False)

    if success_resp is None:
        return {
            'item_num': item_num,
            'found': False,
            'message': '所有候选 OSLC 对象均未成功（见 attempt_log）',
            'attempt_log': attempt_log,
        }

    data = success_resp.json()
    items = data.get('member') or data.get('rdfs:member') or []
    if not items:
        return {
            'item_num': item_num,
            'found': False,
            'message': '查询成功但未找到该物料',
            'used_object': used_object,
            'attempt_log': attempt_log,
        }

    # 标准化（去命名空间前缀）
    raw = items[0]
    clean = {(k.split(':', 1)[1] if ':' in k else k): v for k, v in raw.items()}

    fields_with_values = {k: v for k, v in clean.items() if v is not None and v != ''}
    fields_null = sorted(k for k, v in clean.items() if v is None or v == '')

    FOCUS_FIELDS = [
        'catalogcode', 'newitemdesc', 'manufacturernum', 'modelnum',
        'manufacturer', 'itemnum', 'description', 'itemtype',
        'commodity', 'commoditygroup', 'itemsetid',
    ]
    focus = {f: clean.get(f) for f in FOCUS_FIELDS if f in clean}

    return {
        'item_num': item_num,
        'found': True,
        'used_object': used_object,
        'focus_fields': focus,
        'fields_with_values': fields_with_values,
        'null_fields': fields_null,
        'total_fields': len(clean),
        'attempt_log': attempt_log,
    }


@router.get("/debug/mxitem-for-po/{po_number}", summary="诊断：查看指定PO所有物料的cxmfprodnum/cxtypedsg")
async def debug_mxitem_for_po(po_number: str):
    """
    从 Maximo 抓取指定 PO 的所有 poline，然后批量查询每个 itemnum 的 MXAPIITEM 数据，
    展示 cxmfprodnum（型号）、cxtypedsg（尺寸/类型设计编号）、cxmanufct（制造商）字段值。

    **用途：** 验证 MXAPIITEM 里的 cx 字段是否与 Maximo UI 里的"型号"和"尺寸/质量"列一致。

    操作步骤：
    1. 在 Maximo UI 找一个明显有型号/尺寸数据的 PO（如 CN5074）
    2. 调用此接口，对比返回的 cxmfprodnum / cxtypedsg 与 UI 展示的值是否吻合
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    from src.fetcher.po_fetcher import fetch_po_by_number
    from src.fetcher.item_fetcher import fetch_item_specs

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    try:
        po_data = await loop.run_in_executor(
            executor, lambda: fetch_po_by_number(po_number, save_to_file=False)
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"抓取 Maximo 数据失败: {e}")
    finally:
        executor.shutdown(wait=False)

    if not po_data:
        raise HTTPException(status_code=404, detail=f"未找到 PO: {po_number}（可能认证过期或不存在）")

    poline = po_data.get('poline', [])
    if not poline:
        return {'po_number': po_number, 'poline_count': 0, 'message': '该PO无明细行'}

    # 收集 poline 里物料的原始字段值（catalogcode, newitemdesc）
    poline_summary = []
    item_nums = []
    seen = set()
    for line in poline:
        num = line.get('itemnum')
        if num and num not in seen:
            seen.add(num)
            item_nums.append(num)
        poline_summary.append({
            'polinenum':   line.get('polinenum'),
            'itemnum':     num,
            'description': line.get('description'),
            'catalogcode': line.get('catalogcode'),      # 型号（原始，预期为 null）
            'newitemdesc': line.get('newitemdesc'),      # 尺寸（原始，预期为 null）
            'linetype':    line.get('linetype'),
        })

    # 批量查 MXAPIITEM
    executor2 = ThreadPoolExecutor(max_workers=1)
    try:
        item_spec_map = await loop.run_in_executor(
            executor2, lambda: fetch_item_specs(item_nums)
        )
    except Exception as e:
        item_spec_map = {}
    finally:
        executor2.shutdown(wait=False)

    # 合并结果
    rows = []
    for entry in poline_summary:
        num = entry['itemnum']
        spec = item_spec_map.get(num, {}) if num else {}
        # size_info 解析逻辑（与 map_line_data 保持一致）
        size_info = spec.get('catalogcode')
        if not size_info:
            item_desc = spec.get('description') or ''
            if '/' in item_desc:
                eng_part = item_desc.split('/', 1)[1].strip()
                if '-' in eng_part:
                    size_info = eng_part

        rows.append({
            'polinenum':          entry['polinenum'],
            'itemnum':            num,
            'poline_description': entry['description'],
            # poline 原始值（通常为 null）
            'poline_catalogcode': entry['catalogcode'],
            'poline_newitemdesc': entry['newitemdesc'],
            # MXAPIITEM 字段
            'mxitem_cxtypedsg':   spec.get('cxtypedsg'),    # → model_num（型号）★
            'mxitem_cxmfprodnum': spec.get('cxmfprodnum'),  # 制造商产品编号（参考）
            'mxitem_catalogcode': spec.get('catalogcode'),  # → size_info 首选
            'mxitem_description': spec.get('description'),  # → size_info fallback 来源
            'mxitem_cxmanufct':   spec.get('cxmanufct'),
            # 最终写入 DB 的值（预览）
            'db_model_num':   spec.get('cxtypedsg'),
            'db_size_info':   size_info,
        })

    filled_model = sum(1 for r in rows if r['db_model_num'])
    filled_size  = sum(1 for r in rows if r['db_size_info'])

    return {
        'po_number':     po_number,
        'poline_count':  len(poline),
        'unique_items':  len(item_nums),
        'item_spec_map_hits': len(item_spec_map),
        'filled_model_num': filled_model,
        'filled_size_info': filled_size,
        'rows': rows,
        'hint': (
            '若 mxitem_cxmfprodnum 与 UI 型号列一致、mxitem_cxtypedsg 与 UI 尺寸列一致，'
            '则字段映射正确；否则需调整 ITEM_SPEC_SELECT 中的字段名。'
        ),
    }


@router.get("/debug/po-header-raw/{po_number}", summary="诊断：查看指定PO的供应商/收款方原始字段")
async def debug_po_header_raw(po_number: str):
    """
    从 Maximo MXAPIPO 拉取 PO 主表原始数据，展示供应商字段（ven*）和收款方字段（billto*）。

    用于诊断为何数据库中这些字段为空：
    - 若 Maximo 有值但 DB 为空 → 映射代码或同步需 resync
    - 若 Maximo 本身为空 → 数据问题（Maximo 未填写）
    - 同时展示当前 DB 中该 PO 的对应字段值
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    from config import get_maximo_auth, DEFAULT_HEADERS
    from config.settings import MAXIMO_BASE_URL, VERIFY_SSL
    from config.settings_manager import settings_manager
    import requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    try:
        auth = get_maximo_auth()
    except ValueError as e:
        raise HTTPException(status_code=401, detail=f"认证未配置: {e}")

    headers_http = {
        **DEFAULT_HEADERS,
        'Cookie': auth['cookie'],
        'x-csrf-token': auth['csrf_token'],
    }

    # ── 字段选择（供应商 + 收款方）────────────────────────────────────────
    ven_fields = (
        'ponum,vendor,vendorname,'
        'venaddress1,venaddr1,venaddress2,venaddr2,'
        'venzip,venpostalcode,vencity,venstate,venprovince,'
        'vencontact,venphone,venemail,cxpoemail'
    )
    billto_fields = (
        'billtocomp,billtoname,'
        'billtoaddress1,billtoaddr1,billtoaddress2,billtoaddr2,'
        'billtocity,billtozip,billtopostalcode,billtocountry,'
        'billtocontact,billtophone,billtoemail,shiptoattn'
    )

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    def _fetch():
        url = f"{MAXIMO_BASE_URL}/oslc/os/MXAPIPO"
        params = {
            'oslc.select': f'{ven_fields},{billto_fields}',
            'oslc.where':  f'ponum="{po_number}"',
            '_dropnulls':  '0',
        }
        resp = requests.get(
            url, headers=headers_http, params=params,
            verify=VERIFY_SSL, proxies=settings_manager.get_proxies(), timeout=60,
        )
        return resp

    try:
        resp = await loop.run_in_executor(executor, _fetch)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"请求失败: {e}")

    if resp.status_code == 401:
        raise HTTPException(status_code=401, detail="认证过期，请更新 Cookie")
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Maximo 返回 {resp.status_code}: {resp.text[:300]}")

    data = resp.json()
    members = data.get('member') or data.get('rdfs:member') or []
    if not members:
        raise HTTPException(status_code=404, detail=f"PO {po_number} 未找到")

    # 标准化（去命名空间前缀）
    raw = members[0]
    po = {}
    for k, v in raw.items():
        clean = k.split(':', 1)[1] if ':' in k else k
        po[clean] = v

    # ── 从 DB 取当前已同步值 ──────────────────────────────────────────────
    db_row = None
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """SELECT vendor_code, supplier_name, supplier_address, supplier_address2,
                      supplier_zip, supplier_city, supplier_state,
                      supplier_contact, supplier_phone, supplier_email,
                      company_name, street_address, city, postal_code, country
               FROM purchase_order WHERE code = %s AND del_flag = 0""",
            (po_number,)
        )
        db_row = cursor.fetchone()
        cursor.close()
        conn.close()
    except Exception as e:
        db_row = {'error': str(e)}

    # ── 整理展示 ──────────────────────────────────────────────────────────
    ven_keys = [
        'vendor', 'vendorname',
        'venaddress1', 'venaddr1', 'venaddress2', 'venaddr2',
        'venzip', 'venpostalcode', 'vencity', 'venstate', 'venprovince',
        'vencontact', 'venphone', 'venemail', 'cxpoemail',
    ]
    billto_keys = [
        'billtocomp', 'billtoname',
        'billtoaddress1', 'billtoaddr1', 'billtoaddress2', 'billtoaddr2',
        'billtocity', 'billtozip', 'billtopostalcode', 'billtocountry',
        'billtocontact', 'billtophone', 'billtoemail', 'shiptoattn',
    ]

    maximo_vendor  = {k: po.get(k) for k in ven_keys}
    maximo_billto  = {k: po.get(k) for k in billto_keys}

    filled_ven    = sum(1 for v in maximo_vendor.values() if v)
    filled_billto = sum(1 for v in maximo_billto.values() if v)

    return {
        'po_number':     po_number,
        'maximo_vendor': {
            'filled_count': filled_ven,
            'total_fields': len(ven_keys),
            'fields':       maximo_vendor,
        },
        'maximo_billto': {
            'filled_count': filled_billto,
            'total_fields': len(billto_keys),
            'fields':       maximo_billto,
        },
        'db_current': db_row,
        'hint': (
            'maximo_vendor/maximo_billto 显示 Maximo API 返回值；'
            'db_current 显示当前数据库存储值。'
            '若 Maximo 有值但 DB 为空，执行 POST /api/sync/po/resync 重新同步该 PO。'
            '若 Maximo 返回值也为空，使用 GET /debug/po-all-fields/{po_number} 查看全部字段。'
        ),
    }


@router.get("/debug/po-all-fields/{po_number}", summary="诊断：用*查询PO全部字段，找出供应商/收款方的真实字段名")
async def debug_po_all_fields(po_number: str):
    """
    用 `oslc.select=*` 拉取 PO 全部顶层字段，过滤出含有值的 ven*/billto*/company* 字段。

    **用途：** 当 /debug/po-header-raw 显示 vendorname/venaddress1 等为空时，
    用本接口查出 Maximo 实际使用的供应商/收款方字段名（可能因版本不同而异）。
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    from config import get_maximo_auth, DEFAULT_HEADERS
    from config.settings import MAXIMO_BASE_URL, VERIFY_SSL
    from config.settings_manager import settings_manager
    import requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    try:
        auth = get_maximo_auth()
    except ValueError as e:
        raise HTTPException(status_code=401, detail=f"认证未配置: {e}")

    headers_http = {
        **DEFAULT_HEADERS,
        'Cookie': auth['cookie'],
        'x-csrf-token': auth['csrf_token'],
    }

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    def _fetch():
        url = f"{MAXIMO_BASE_URL}/oslc/os/MXAPIPO"
        params = {
            'oslc.select': '*',
            'oslc.where':  f'ponum="{po_number}"',
            '_dropnulls':  '0',
        }
        return requests.get(
            url, headers=headers_http, params=params,
            verify=VERIFY_SSL, proxies=settings_manager.get_proxies(), timeout=60,
        )

    try:
        resp = await loop.run_in_executor(executor, _fetch)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"请求失败: {e}")

    if resp.status_code == 401:
        raise HTTPException(status_code=401, detail="认证过期")
    if resp.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Maximo {resp.status_code}: {resp.text[:300]}")

    data = resp.json()
    members = data.get('member') or data.get('rdfs:member') or []
    if not members:
        raise HTTPException(status_code=404, detail=f"PO {po_number} 未找到")

    # 标准化（去命名空间前缀）
    raw = members[0]
    po: dict = {}
    for k, v in raw.items():
        clean = k.split(':', 1)[1] if ':' in k else k
        po[clean] = v

    # ── 过滤出供应商/收款方相关字段（含值的）────────────────────────────
    VENDOR_PREFIXES  = ('ven', 'vendor', 'company', 'supp')
    BILLTO_PREFIXES  = ('billto', 'shipto', 'shiptoattn')
    OTHER_KEYS       = ('ponum', 'description', 'status')

    vendor_fields  = {}
    billto_fields  = {}
    other_fields   = {}
    all_keys_nonempty = {}

    for k, v in po.items():
        if isinstance(v, (dict, list)):
            continue  # 跳过子资源
        lo = k.lower()
        if any(lo.startswith(p) for p in VENDOR_PREFIXES):
            vendor_fields[k] = v
        elif any(lo.startswith(p) for p in BILLTO_PREFIXES):
            billto_fields[k] = v
        elif k in OTHER_KEYS:
            other_fields[k] = v
        # 收集所有有值的顶层字段（供参考）
        if v not in (None, '', 0, False):
            all_keys_nonempty[k] = v

    return {
        'po_number':     po_number,
        'total_fields_returned': len(po),
        'vendor_fields':  vendor_fields,
        'billto_fields':  billto_fields,
        'other_fields':   other_fields,
        'all_nonempty_scalar_fields': all_keys_nonempty,
        'hint': (
            '查看 vendor_fields 找供应商名称/地址的真实字段名，'
            '查看 billto_fields 找收款方信息的真实字段名。'
            '将找到的字段名更新到 src/utils/mapper.py 的 VENDOR_FIELD_CANDIDATES / BILLTO_FIELD_CANDIDATES。'
        ),
    }


@router.get("/debug/company/{company_code}", summary="诊断：从MXAPICOMPANY查询公司（供应商/收款方）详细信息")
async def debug_company(
    company_code: str,
    api_object: str = Query("MXAPICOMPANY", description="Maximo API 对象名，如 MXAPICOMPANY / MXAPIVENDOR"),
):
    """
    查询 Maximo 公司档案 API，用于获取供应商/收款方地址信息。

    **用法：**
    - `GET /debug/company/8970301` — 查询供应商 8970301 的公司信息
    - `GET /debug/company/BILLTOCHINA` — 查询收款方 BILLTOCHINA 的公司信息
    - `GET /debug/company/8970301?api_object=MXAPIVENDOR` — 尝试 MXAPIVENDOR

    返回 Maximo 公司档案的全部标量字段（含名称、地址等）。
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    from config import get_maximo_auth, DEFAULT_HEADERS
    from config.settings import MAXIMO_BASE_URL, VERIFY_SSL
    from config.settings_manager import settings_manager
    import requests, urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    try:
        auth = get_maximo_auth()
    except ValueError as e:
        raise HTTPException(status_code=401, detail=f"认证未配置: {e}")

    headers_http = {
        **DEFAULT_HEADERS,
        'Cookie': auth['cookie'],
        'x-csrf-token': auth['csrf_token'],
    }

    loop = asyncio.get_event_loop()
    executor = ThreadPoolExecutor(max_workers=1)

    # 尝试多种 where 条件（不同 Maximo 版本字段名不同）
    where_attempts = [
        f'company="{company_code}"',
        f'vendor="{company_code}"',
        f'companynum="{company_code}"',
    ]

    def _fetch(where_clause: str):
        url = f"{MAXIMO_BASE_URL}/oslc/os/{api_object}"
        params = {
            'oslc.select': '*',
            'oslc.where':  where_clause,
            '_dropnulls':  '0',
            'oslc.pageSize': 1,
        }
        return requests.get(
            url, headers=headers_http, params=params,
            verify=VERIFY_SSL, proxies=settings_manager.get_proxies(), timeout=60,
        )

    results = []
    for where in where_attempts:
        try:
            resp = await loop.run_in_executor(executor, lambda w=where: _fetch(w))
            if resp.status_code == 200:
                data = resp.json()
                members = data.get('member') or data.get('rdfs:member') or []
                if members:
                    raw = members[0]
                    record = {}
                    for k, v in raw.items():
                        clean = k.split(':', 1)[1] if ':' in k else k
                        if not isinstance(v, (dict, list)):
                            record[clean] = v
                    results.append({'where': where, 'found': True, 'fields': record})
                    break  # 找到就停止
                else:
                    results.append({'where': where, 'found': False, 'http_status': resp.status_code})
            else:
                results.append({'where': where, 'found': False, 'http_status': resp.status_code,
                                'error': resp.text[:200]})
        except Exception as e:
            results.append({'where': where, 'found': False, 'error': str(e)})

    found = next((r for r in results if r.get('found')), None)
    return {
        'company_code': company_code,
        'api_object':   api_object,
        'attempts':     results,
        'company_fields': found['fields'] if found else None,
        'hint': (
            '若 company_fields 中有名称/地址字段，说明可通过此 API 获取供应商/收款方信息。'
            '若所有 where 条件都返回 found=false，尝试参数 api_object=MXAPIVENDOR 或 MXAPIADDRESS。'
        ),
    }
